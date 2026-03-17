import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side

def format_report(report_path, config):

    data_frame = pd.read_excel(report_path)

    # Filters out the columns not required and then custom sorts the remaining rows
    data_frame = data_frame[config["columns_needed"]]
    
    # Ensures time column is sorted by datetime values, not strings
    data_frame["Time"] = pd.to_datetime(data_frame["Time"], format="mixed")

    data_frame = data_frame.sort_values(
        by=config["sort_by"],
        ascending=[True] * len(config["sort_by"])
    )

    data_frame["Time"] = data_frame["Time"].dt.strftime("%I:%M%p")

    # Creates a new .xlsx file and adds data from the data frame
    output_folder = config["output_folder"]
    os.makedirs(output_folder, exist_ok = True)
    output_file = os.path.join(output_folder, config["output_filename"])
    data_frame.to_excel(output_file, index = False)

    workbook = load_workbook(output_file)
    worksheet = workbook.active
    apply_styling(worksheet, config)

    # Landscape printing
    worksheet.page_setup.orientation = worksheet.ORIENTATION_LANDSCAPE

    workbook.save(output_file)
    return output_file

def apply_styling(worksheet, config):

    column_widths = {
        "A": 20,  # Class
        "B": 11,  # Day
        "C": 9,  # Time
        "D": 18,  # Zone
        "E": 16,  # Class Trainer
        "F": 18,  # Student User Number
        "G": 15,  # Student Name
        "H": 15,  # Student Surname
    }

    # Changes the width of each column to match the correct sizing of its header
    for column, width in column_widths.items():
        worksheet.column_dimensions[column].width = width

    # Defines the border styling
    thin = Side(border_style = "thin", color = "000000")
    border = Border(top = thin, left = thin, right = thin, bottom = thin)

    # Applies a thin border to all cells in the worksheet
    for row in worksheet.iter_rows():
        for cell in row:
            cell.border = border

    # Finds at which position the "Class Trainer" column is found
    header_row = next(worksheet.iter_rows(min_row = 1, max_row = 1))
    column_index = next((i for i, cell in enumerate(header_row) if cell.value == "Class Trainer"), None)

    if column_index is None:
        raise ValueError("Column 'Class Trainer' not found in header row")

    # Fills in the rows with alternating colours, switching when the trainer changes
    colours = config["trainer_colours"]
    colour_index = 1
    prev_trainer = None

    # Applies a light grey fill to the cells in the first row (header row)
    for cell in worksheet[1]:
        cell.fill = PatternFill(patternType = "solid", fgColor = colours[2])

    for row in worksheet.iter_rows(min_row = 2):
        trainer_curr = row[column_index].value

        if trainer_curr != prev_trainer:
            colour_index = 1 - colour_index
            prev_trainer = trainer_curr

        for cell in row:
            cell.fill = PatternFill(patternType = "solid", fgColor = colours[colour_index])
